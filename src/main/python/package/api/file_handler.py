from openpyxl import load_workbook


def load_wb(name):
    # check if the file is an excel file
    if name.endswith('.xlsx'):
        return load_workbook(name)
    else:
        raise Exception('File is not an excel file')

def get_col_names(sheet):
    max_col = sheet.max_column
    col_names = []
    for i in range(1, max_col + 1):
        cell_obj = sheet.cell(row=1, column=i)
        col_names.append({'letter' : cell_obj.column_letter, 'value': cell_obj.value})

    return col_names
